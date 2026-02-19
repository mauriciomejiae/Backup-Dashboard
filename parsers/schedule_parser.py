"""Parser del archivo Excel de Schedule mensual."""

import re
import openpyxl
from models.report_data import ScheduleRow, ScheduleReport


# Mapeo de hojas del Schedule a nombres de Cell Manager
SHEET_MAPPING = {
    "COMHP81": "COMHP81",
    "COMHP83": "COMHP83",
    "LNXCELLMNGVEN": "LNXCELLMNGVEN",
    "LNXCELLMNGTRI": "LNXCELLMNGTRI",
    "LNXCELLMNGPTA": "LNXCELLMNGPTA",
    "NETBACKUP": "NETBACKUP",
    "COMMVAULT_NBUIT": "COMMVAULT_NBUIT",
    "COMMVAULT_OCI": "COMMVAULT_OCI",
    # "ACRONIS": "ACRONIS",  # Excluido explícitamente
}

# Patrones de tickets ITSM
ITSM_PATTERN = re.compile(r"^(WO|RF|CHG|REQ|INC)", re.IGNORECASE)


def _is_itsm_ticket(value) -> bool:
    """Verifica si un valor es un ticket ITSM (WOxxx, RFxxx, CHGxxx, REQxxx, INCxxx)."""
    if value is None:
        return False
    text = str(value).strip()
    if not text:
        return False
    return bool(ITSM_PATTERN.match(text))


def parse_schedule_sheet(ws, sheet_name: str) -> dict:
    """Parsea una hoja del Schedule y cuenta estados.

    Retorna dict con conteos: ejecutados, programados, relanzados, fallidos, q (casos ITSM).
    """
    total_programados = 0
    total_ejecutados = 0
    total_fallidos = 0
    total_relanzados = 0
    total_gestionados = 0
    total_q = 0  # Casos ITSM creados

    # Leer headers
    headers = []
    status_col = None
    job_id_relanzado_col = None
    caso_col = None

    for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]:
        val = str(cell.value).strip().upper() if cell.value else ""
        headers.append(val)

    # Encontrar columnas relevantes
    for i, h in enumerate(headers):
        if "STATUS" in h or "ESTADO" in h:
            status_col = i
        if "RELANZADO" in h or "JOB ID RELANZADO" in h or "JOBID RELANZADO" in h:
            job_id_relanzado_col = i
        if "CASO" in h or "TICKET" in h or "ITSM" in h:
            caso_col = i

    if status_col is None:
        status_col = 3  # Columna D por defecto

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue

        total_programados += 1
        
        # Obtener status y normalizar
        status_val = str(row[status_col]).strip().lower() if len(row) > status_col and row[status_col] else ""
        
        # 1. Determinar si es Fallido (Regla estricta: Failed o Aborted)
        is_failed = status_val in ("failed", "aborted")
        if is_failed:
            total_fallidos += 1

        # 2. Determinar si es Relanzado (Status 'relaunched' O columna Relanzado con valor)
        # Verificar ID de relanzamiento
        has_relaunch_id = False
        if job_id_relanzado_col is not None and len(row) > job_id_relanzado_col:
             relanzado_val = row[job_id_relanzado_col]
             if relanzado_val is not None and str(relanzado_val).strip() not in ("", "None", "nan"):
                 has_relaunch_id = True

        is_relaunched = False
        if "relaunched" in status_val:
            is_relaunched = True
        elif has_relaunch_id:
            is_relaunched = True
        
        if is_relaunched:
            total_relanzados += 1

        # 3. Determinar Ejecutados
        # Si no es fallido, asumimos que fue ejecutado (Completed, Warning, Relaunched, Success, etc.)
        # A menos que esté vacío completamente
        if status_val and not is_failed:
            total_ejecutados += 1

        # 4. Contar caso ITSM (Q) - Independiente del status
        has_case = False
        if caso_col is not None and len(row) > caso_col:
            if _is_itsm_ticket(row[caso_col]):
                total_q += 1
                has_case = True
        elif caso_col is None:
            # Si no hay columna CASO explícita, buscar en todas las celdas
            for cell_val in row:
                if _is_itsm_ticket(cell_val):
                    total_q += 1
                    has_case = True
                    break  # Un caso por fila máximo

        # 5. Determinar Fallidos Gestionados
        # Regla: (Failed/Aborted Y con Caso) O (Relanzado Y (con Caso O con ID Relanzamiento))
        is_managed = False
        
        if is_failed and has_case:
            is_managed = True
        elif is_relaunched:
            # Si es relanzado (ya validado por status o ID), chequeamos gestión
            if has_case or has_relaunch_id:
                is_managed = True
        
        if is_managed:
            total_gestionados += 1

    return {
        "programados": total_programados,
        "ejecutados": total_ejecutados,
        "fallidos": total_fallidos,
        "relanzados": total_relanzados,
        "gestionados": total_gestionados,
        "q": total_q,
    }


def parse_schedule_file(file_path: str, period_name: str = "") -> ScheduleReport:
    """Parsea el archivo Excel del Schedule mensual completo.

    Lee cada hoja mapeada y genera el ScheduleReport con KPIs.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    rows = []

    for sheet_name, platform_name in SHEET_MAPPING.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            data = parse_schedule_sheet(ws, sheet_name)

            programados = data["programados"]
            ejecutados = data["ejecutados"]
            fallidos = data["fallidos"]
            relanzados = data["relanzados"]
            q = data["q"]
            gestionados = data["gestionados"]

            # KPIs
            kpi_op = ((programados - fallidos) / programados * 100) if programados > 0 else 0
            pct_rel = (relanzados / (programados - fallidos) * 100) if (programados - fallidos) > 0 else 0
            # Denominador incluye fallidos y relanzados
            gestion_f = (gestionados / (fallidos + relanzados) * 100) if (fallidos + relanzados) > 0 else 0

            row = ScheduleRow(
                platform=platform_name,
                ejecutados=ejecutados,
                programados=programados,
                relanzados=relanzados,
                fallidos=fallidos,
                q=q,
                gestionados=gestionados,
                kpi_operacion=round(kpi_op, 2),
                pct_relanzamiento=round(pct_rel, 2),
                gestion_fallidos=round(gestion_f, 2),
            )
            rows.append(row)

    # Totales
    total_ej = sum(r.ejecutados for r in rows)
    total_prog = sum(r.programados for r in rows)
    total_rel = sum(r.relanzados for r in rows)
    total_fal = sum(r.fallidos for r in rows)
    total_q = sum(r.q for r in rows)
    total_gest = sum(r.gestionados for r in rows)

    kpi_op_gen = ((total_prog - total_fal) / total_prog * 100) if total_prog > 0 else 0
    kpi_gest_gen = (total_gest / (total_fal + total_rel) * 100) if (total_fal + total_rel) > 0 else 0
    pct_rel_gen = (total_rel / (total_prog - total_fal) * 100) if (total_prog - total_fal) > 0 else 0

    report = ScheduleReport(
        period_name=period_name,
        rows=rows,
        total_ejecutados=total_ej,
        total_programados=total_prog,
        total_relanzados=total_rel,
        total_fallidos=total_fal,
        total_q=total_q,
        kpi_operacion_general=round(kpi_op_gen, 2),
        kpi_gestion_fallidos_general=round(kpi_gest_gen, 2),
        pct_relanzados_general=round(pct_rel_gen, 2),
    )

    wb.close()
    return report
